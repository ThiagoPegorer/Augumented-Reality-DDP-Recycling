"""
GUI Monte Carlo parser - replaces the IPC MC runner (which hit the ~300-draw session
leak; the GUI simulator does not). Reads the openLCA GUI simulation exports and rebuilds
the SAME output files as mc_run.py, so everything downstream is unchanged.

Input (Outputs\4_monte_carlo\): one Excel export per product system, named
    simulation_result_Sc1.xlsx   simulation_result_Sc2g.xlsx  simulation_result_Sc2s.xlsx
    simulation_result_Sc3g.xlsx  simulation_result_Sc3s.xlsx  simulation_result_Sc4g.xlsx
    simulation_result_Sc4s.xlsx
(GUI: right-click product system -> Calculate -> Monte Carlo Simulation, 1000 runs,
 EF 3.1 Method (adapted) -> Export -> save with the name above.)

Safety: each file's 'Calculation setup' sheet is checked against the expected product-
system name - a mislabeled export (e.g. Sc3g saved as Sc3s) is REJECTED, not parsed.

Output (Outputs\4_monte_carlo\): mc_raw_<key>.csv (overwritten from the GUI draws), mc_summary.csv,
mc_net.csv, mc_gui_parse_log.txt. Missing exports are skipped with a note, so you can
run this after every few GUI simulations to see intermediate state.

One-time dependency:  py -m pip install openpyxl
Run:  py "C:\Claude\Projects\AR_DPP\LCA_Analysis\Scripts\4_monte_carlo\mc_gui_parse.py"
"""
import csv, math, pathlib, traceback

HERE = pathlib.Path(__file__).resolve().parent.parent.parent / "Outputs" / pathlib.Path(__file__).resolve().parent.name
HERE.mkdir(parents=True, exist_ok=True)
TXT = HERE / "mc_gui_parse_log.txt"; _l = []
def log(m):
    print(m); _l.append(str(m))
    try: TXT.write_text("\n".join(_l), encoding="utf-8")
    except OSError: pass

SYSTEMS = [   # (key, exact product-system name for the safety check)
    ("Sc1",  "VCU S5 EoL Sc1 (no recycling)"),
    ("Sc2g", "VCU S5 EoL Sc2 (bulk recycling)"),
    ("Sc2s", "VCU S5 EoL Sc2 credits (avoided virgin)"),
    ("Sc3g", "VCU S5 EoL Sc3 (guided disassembly)"),
    ("Sc3s", "VCU S5 EoL Sc3 credits (avoided virgin)"),
    ("Sc4g", "VCU S5 EoL Sc4 (disassembly + reuse)"),
    ("Sc4s", "VCU S5 EoL Sc4 credits (avoided virgin + components)"),
]
NETS = [("Sc2", "Sc2g", "Sc2s"), ("Sc3", "Sc3g", "Sc3s"), ("Sc4", "Sc4g", "Sc4s")]

def percentile(sorted_xs, q):
    if not sorted_xs: return float("nan")
    k = (len(sorted_xs) - 1) * q / 100.0
    f, c = int(math.floor(k)), int(math.ceil(k))
    if f == c: return sorted_xs[f]
    return sorted_xs[f] + (sorted_xs[c] - sorted_xs[f]) * (k - f)

def stats(xs):
    n = len(xs)
    if n == 0: return None
    mean = sum(xs) / n
    sd = math.sqrt(sum((x - mean) ** 2 for x in xs) / (n - 1)) if n > 1 else 0.0
    s = sorted(xs)
    return [n, mean, sd] + [percentile(s, q) for q in (5, 25, 50, 75, 95)]

def parse_export(path, expected_name):
    """Return (cats, units, rows) - rows: one list per iteration, ordered like cats."""
    import openpyxl, warnings
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # ---- safety check: the export really is the expected product system ----
    setup = wb["Calculation setup"]
    ps_name = None
    for row in setup.iter_rows(max_row=15, values_only=True):
        vals = [v for v in row if v is not None]
        if len(vals) >= 2 and str(vals[0]).startswith("Product system"):
            ps_name = str(vals[1]).strip(); break
    if ps_name != expected_name:
        raise ValueError(f"export is for '{ps_name}', expected '{expected_name}' - "
                         f"wrong file name? NOT parsed.")

    ws = wb["Impact Assessment"]
    it = ws.iter_rows(values_only=True)
    header = None
    for row in it:                                  # find the header row
        if row and "Impact category" in [str(v) for v in row if v is not None]:
            header = list(row); break
    if header is None: raise ValueError("no 'Impact category' header row found")
    i_cat  = header.index("Impact category")
    i_unit = header.index("Reference unit")
    i_run1 = next(i for i, v in enumerate(header) if str(v).strip() == "Run 1")

    cats, units, per_cat = [], {}, {}
    for row in it:
        if not row or row[i_cat] is None: continue
        c = str(row[i_cat]).strip()
        cats.append(c)
        units[c] = str(row[i_unit] or "").strip()
        per_cat[c] = [float(v) for v in row[i_run1:] if v is not None]
    wb.close()

    ns = {len(v) for v in per_cat.values()}
    if len(ns) != 1:
        raise ValueError(f"inconsistent run counts across categories: {sorted(ns)}")
    n = ns.pop()
    cats = sorted(cats, key=str.lower)              # same column order as mc_run.py
    rows = [[per_cat[c][i] for c in cats] for i in range(n)]
    return cats, units, rows, n

def main():
    log("Parsing GUI Monte Carlo exports ...")
    data = {}
    for skey, name in SYSTEMS:
        path = HERE / f"simulation_result_{skey}.xlsx"
        if not path.exists():
            log(f"  {skey}: {path.name} not found - skipped."); continue
        try:
            cats, units, rows, n = parse_export(path, name)
        except Exception as ex:
            log(f"  !! {skey}: {ex}"); continue
        data[skey] = (cats, units, rows)
        with (HERE / f"mc_raw_{skey}.csv").open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["iteration"] + cats)
            for i, r in enumerate(rows, 1): w.writerow([i] + r)
        log(f"  {skey}: {n} draws x {len(cats)} categories -> mc_raw_{skey}.csv")

    if not data:
        log("Nothing parsed - no output rebuilt."); return

    log("\nWriting mc_summary.csv ...")
    with (HERE / "mc_summary.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["system", "category", "unit", "n", "mean", "sd",
                    "p5", "p25", "p50", "p75", "p95"])
        for skey, _ in SYSTEMS:
            if skey not in data: continue
            cats, units, rows = data[skey]
            for j, c in enumerate(cats):
                st = stats([r[j] for r in rows])
                if st: w.writerow([skey, c, units.get(c, "")] + st)

    log("Writing mc_net.csv (net = gross - saving, run-paired, independent-sampling "
        "approximation) ...")
    with (HERE / "mc_net.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["scenario", "category", "n_pairs", "mean", "sd",
                    "p5", "p25", "p50", "p75", "p95"])
        for scen, gk, sk in NETS:
            if gk not in data or sk not in data:
                log(f"  {scen}: missing {gk if gk not in data else sk} export - skipped.")
                continue
            gc, _, gr = data[gk]
            sc_, _, sr = data[sk]
            npairs = min(len(gr), len(sr))
            for j, c in enumerate(gc):
                if c not in sc_: continue
                js = sc_.index(c)
                st = stats([gr[i][j] - sr[i][js] for i in range(npairs)])
                if st: w.writerow([scen, c] + st)

    done = [k for k, _ in SYSTEMS if k in data]
    missing = [k for k, _ in SYSTEMS if k not in data]
    log(f"\nDONE. Parsed: {', '.join(done)}."
        + (f" Still missing: {', '.join(missing)}." if missing else " All 7 complete."))
    log("Send mc_gui_parse_log.txt + mc_summary.csv + mc_net.csv.")

if __name__ == "__main__":
    try: main()
    except Exception: log("FATAL:\n" + traceback.format_exc())
    finally:
        try: TXT.write_text("\n".join(_l), encoding="utf-8")
        except OSError: pass
        print(f"\nLog: {TXT}")
